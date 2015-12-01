from __future__ import absolute_import
import six
import warnings
from .base import ExcelProducerBase
from openpyxl import Workbook
from openpyxl.writer.write_only import WriteOnlyCell
from openpyxl.styles import colors, Font, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


class OpenpyxlProducer(ExcelProducerBase):

    width_factor = 1.04
    default_extension = '.xlsx'
    default_windows_links = True

    def __init__(self, *args, **kwargs):
        super(OpenpyxlProducer, self).__init__(*args, **kwargs)

    def create_book(self, filepath):
        book = Workbook(optimized_write=self.optimize)
        if not self.optimize:
            book.remove_sheet(book.active)
        return book

    def close_book(self, book, filepath):
        book.save(filepath)

    def create_sheet(self, book, shname):
        return book.create_sheet(shname or 'Sheet1')

    def close_sheet(self, book, sheet, last_row, last_col):
        if self.optimize:
            if self.show_warnings:
                warnings.warn('cannot freeze panes in optimized mode')
        else:
            sheet.freeze_panes = sheet.cell(row=2, column=min(last_col, 1) + 1)
        if last_row > 0 and last_col >= 0:
            sheet.auto_filter.ref = 'A1:%s%d' % (
                get_column_letter(last_col + 1), last_row + 1)
        if last_row > 1 and last_col > 1:
            if self.optimize and self.show_warnings:
                warnings.warn('cannot hide grid lines in optimized mode')
            sheet.sheet_view.showGridLines = False

    def set_dimensions(self, book, sheet, fields, format):
        for c, f in enumerate(fields, start=1):
            w = format[f].get('width')
            if w:
                col = get_column_letter(c)
                sheet.column_dimensions[col].width = w * self.width_factor

    def make_styles(self, book):
        self.font_head = Font(bold=True)
        self.align_head = Alignment(horizontal='center')
        self.align_left = Alignment(horizontal='left', vertical='top')
        self.align_center = Alignment(horizontal='center', vertical='top')
        self.align_wrap = Alignment(horizontal='left', vertical='top',
                                    wrap_text=True)
        self.font_link = Font(color=colors.BLUE, underline='single')
        side_light = Side(border_style='thin', color='FF999999')
        side_dark = Side(border_style='medium', color='FF666666')
        self.border_head = Border(right=side_light, bottom=side_dark)
        self.border_data = Border(right=side_light, bottom=side_light)

    def make_header(self, book, sheet, fields, format):
        row = 1
        row_cells = []
        optimize = self.optimize

        for col, f in enumerate(fields, start=1):
            if optimize:
                cell = WriteOnlyCell(sheet)
            else:
                cell = sheet.cell(row=row, column=col)
            cell.set_explicit_value(self.get_col_name(f))
            cell.font = self.font_head
            cell.alignment = self.align_head
            cell.border = self.border_head
            if optimize:
                row_cells.append(cell)
        if optimize:
            sheet.append(row_cells)

    def data_row(self, row, item, book, sheet, abs_row, fields, format):
        row += 1
        row_cells = []
        img_no = 1
        optimize = self.optimize

        for col, f in enumerate(fields, start=1):
            align = format[f]['align']
            type_ = format[f]['type']
            value, link = self.get_value_link(item, f, shorten=True)
            text = six.text_type(value)
            img_path = self.get_image_path(item, f)

            if img_path:
                img_val = Image(img_path)
                if self.image_shift is None:
                    img_col = len(fields) + img_no
                    cell = sheet.cell(row=row, column=img_col)
                    cell.value = self.blank_value
                else:
                    img_col = col + self.image_shift
                anchor = '%s%d' % (get_column_letter(img_col), row)
                sheet.add_image(img_val, anchor)
                img_no += 1

            if optimize:
                cell = WriteOnlyCell(sheet)
            else:
                cell = sheet.cell(row=row, column=col)

            if link:
                cell.set_explicit_value(self.maybe_blank(text))
                cell.hyperlink = link
            elif type_ == 'string' or link is False or text == '':
                cell.set_explicit_value(
                    self.maybe_blank(text), cell.TYPE_STRING)
            elif type_ in ('int', 'float', 'number', 'currency'):
                number = self.as_number(value)
                if number is None:
                    cell.set_explicit_value(
                        self.maybe_blank(text), cell.TYPE_STRING)
                else:
                    cell.set_explicit_value(number, cell.TYPE_NUMERIC)
                    if type_ == 'currency':
                        cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            else:
                cell.value = self.maybe_blank(value)

            if link:
                cell.font = self.font_link
            if align == 'wrap':
                cell.alignment = self.align_wrap
            elif align == 'center':
                cell.alignment = self.align_center
            else:
                cell.alignment = self.align_left
            cell.border = self.border_data

            if optimize:
                row_cells.append(cell)

        if self.hard_blanks:
            if optimize:
                row_cells.append(WriteOnlyCell(sheet, value=self.blank_value))
            else:
                sheet.cell(row=row, column=col + 1).value = self.blank_value

        if optimize:
            sheet.append(row_cells)
